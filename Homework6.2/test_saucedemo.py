import pytest
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from webdriver_manager.chrome import ChromeDriverManager
from pathlib import Path
from datetime import date
from constants import globalConstants
import openpyxl 
from data import valid_login
from data import locked_login


class Test_Saucedemo:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.pathself = Path(__file__).parent.resolve()
        self.datetime = date.today().strftime('%d-%m-%Y')

        WebDriverWait(self.driver, 0.5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"username\"]")))
        WebDriverWait(self.driver, 0.5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"password\"]")))
        self.usernameInput = self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"username\"]")
        self.passwordInput = self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"password\"]")
        WebDriverWait(self.driver, 0.5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"login-button\"]")))
        self.loginBttn = self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"login-button\"]")

    def teardown_method(self):
        self.driver.quit()

    def getInvalidData():
        excelFile = openpyxl.load_workbook("C:\\Users\\Ali\\git\\PythonSeleniumCamp2023\\PythonSeleniumCamp2023\\Homework6.2\\data\\invalid_logins.xlsx")
        selectedSheet = excelFile["Sayfa1"]

        totalRows = selectedSheet.max_row
        data =[]
        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        
        return data

    def login(self):
        self.usernameInput.click()
        self.usernameInput.send_keys(valid_login.validUsername)
        self.passwordInput.click()
        self.passwordInput.send_keys(valid_login.validPassword)
        self.loginBttn.click()


    def test_validLogin(self):
        self.usernameInput.click()
        self.usernameInput.send_keys(valid_login.validUsername)
        self.passwordInput.click()
        self.passwordInput.send_keys(valid_login.validPassword)
        self.loginBttn.click()
        WebDriverWait(self.driver, 0.5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, "#item_4_img_link > .inventory_item_img")))
        self.driver.save_screenshot(f"{self.pathself}/test-validLogin-{valid_login.validUsername}-{valid_login.validPassword}-{self.datetime}.png")
        assert self.driver.current_url == globalConstants.loginURL

    @pytest.mark.parametrize("username,password",getInvalidData())
    def test_invalidLogin(self,username,password):
        self.usernameInput.send_keys(username)
        self.passwordInput.send_keys(password)
        self.loginBttn.click()
        errorMessage = self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"error\"]")
        self.driver.save_screenshot(f"{self.pathself}/test-invalidLogin-{username}-{password}-{self.datetime}.png")
        assert errorMessage.text == globalConstants.invalidLoginErrorText

    def test_emptyAreas(self):
        self.loginBttn.click()
        error = self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"error\"]")
        WebDriverWait(self.driver, 0.5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"error\"]")))
        self.driver.save_screenshot(f"{self.pathself}/test-emptyAreas-{self.datetime}.png")
        assert error.text == globalConstants.emptyLoginErrorText

    def test_emptyPassword(self):
        self.usernameInput.send_keys(valid_login.validUsername)
        self.loginBttn.click()
        error = self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"error\"]")
        self.driver.save_screenshot(f"{self.pathself}/test-emptyPassword-{self.datetime}.png")
        assert error.text == globalConstants.emptyPasswordText

    def test_xbttnsRemoved(self):
        self.loginBttn.click()
        alertBttn = self.driver.find_element(By.CSS_SELECTOR, ".fa-times > path")
        alertBttn.click()
        xIconUp = self.driver.find_elements(By.CSS_SELECTOR, ".form_group:nth-child(1) path")
        xIconDown = self.driver.find_elements(By.CSS_SELECTOR, ".form_group:nth-child(2) > .svg-inline--fa")
        self.driver.save_screenshot(f"{self.pathself}/test-xbttnsRemoved-{self.datetime}.png")
        assert (len(xIconDown) & len(xIconUp) == 0)

    def test_addCardBttn(self):
        self.login()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-sauce-labs-backpack\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-sauce-labs-bike-light\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-sauce-labs-bolt-t-shirt\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-sauce-labs-fleece-jacket\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-sauce-labs-onesie\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-test.allthethings()-t-shirt-(red)\"]").click()

        self.driver.save_screenshot(f"{self.pathself}/test-addCardBttn-{self.datetime}.png")

        assert self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"remove-sauce-labs-backpack\"]").text == "Remove"
        assert self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"remove-sauce-labs-bike-light\"]").text == "Remove"
        assert self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"remove-sauce-labs-bolt-t-shirt\"]").text == "Remove"
        assert self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"remove-sauce-labs-fleece-jacket\"]").text == "Remove"
        assert self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"remove-sauce-labs-onesie\"]").text == "Remove"
        assert self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"remove-test.allthethings()-t-shirt-(red)\"]").text == "Remove"

    def test_itemNumbers(self):
        self.login()
        itemNumbers = self.driver.find_elements(By.CLASS_NAME,"inventory_item")
        self.driver.save_screenshot(f"{self.pathself}/test-itemNumbers-{self.datetime}.png")
        assert len(itemNumbers)==6

    def test_lockedUser(self):
        self.usernameInput.send_keys(locked_login.lockedUsername)
        self.passwordInput.send_keys(locked_login.lockedPassword)
        self.loginBttn.click()
        errorMessage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.pathself}/test-lockedUser-{locked_login.lockedUsername}-{locked_login.lockedPassword}-{self.datetime}.png")
        assert errorMessage.text == globalConstants.lockedUserText

    def test_logOutBttn(self):
        self.login()
        burgerMenu = self.driver.find_element(By.ID, "react-burger-menu-btn")
        burgerMenu.click()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.ID, "logout_sidebar_link")))
        logOutBttn = self.driver.find_element(By.ID, "logout_sidebar_link")
        logOutBttn.click()
        self.driver.save_screenshot(f"{self.pathself}/test-logOutBttn-{self.datetime}.png")
        assert self.driver.current_url==globalConstants.URL
    
    def test_shopIconText(self):
        self.login()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-sauce-labs-backpack\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-sauce-labs-bike-light\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-sauce-labs-bolt-t-shirt\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-sauce-labs-fleece-jacket\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-sauce-labs-onesie\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"add-to-cart-test.allthethings()-t-shirt-(red)\"]").click()
        self.driver.save_screenshot(f"{self.pathself}/test-shopIconText-{self.datetime}.png")
        assert self.driver.find_element(By.CSS_SELECTOR, ".shopping_cart_badge").text == "6"


   

    



    


    